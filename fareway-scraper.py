from email.mime.text import MIMEText
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import WebDriverException, ElementClickInterceptedException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib, ssl



import datetime
import time





MAX_WAIT = 30
NOW = datetime.datetime.now()
STORES_TO_SCRAPE = set(['Atlantic', 'Carroll', 'Clarinda', 'Council Bluffs', 'Creston',
                        'Denison', 'Harlan', 'Indianola', 'Jefferson', 'Osceola',
                        'Red Oak', 'Shenandoah', 'Winterset'])
TODAY = datetime.date.today()

if NOW.hour >= 15 and NOW.minute >= 45:
    DAY = datetime.date(TODAY.year, TODAY.month, TODAY.day + 1) 
elif NOW.hour >= 16:
    DAY = datetime.date(TODAY.year, TODAY.month, TODAY.day + 1) 
else:
    DAY = TODAY


with webdriver.Firefox() as browser:

    def wait_for_javascript():
        start_time = time.time()
        while True:
            try:
                element = WebDriverWait(browser, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//button[@click.delegate='selectStore(store, $event)']")))
                print("found header")
                
                return
            except (AssertionError, WebDriverException, IndexError) as e:
                if time.time() - start_time > MAX_WAIT:
                    raise e
                time.sleep(0.5)

    def wait_for_store_list(storeName):
        start_time = time.time()
        while True:
            try:
                stores = browser.find_elements(By.XPATH,  "//span[@class='store-select-store-name']")
                storeText = [store.text for store in stores]
                assert(storeName in storeText)
                print("found store in store")
                return
            except (AssertionError, WebDriverException, IndexError) as e:
                if time.time() - start_time > MAX_WAIT:
                    browser.find_element(By.XPATH, "//a[@click.delegate='showFulfillmentTimes($event)']").click()
                time.sleep(0.5)

    browser.get("https://shop.fareway.com")
    wait_for_javascript()
    stores = browser.find_elements(By.XPATH,  "//span[@class='store-select-store-name']")
    storeText = [store.text for store in stores]
    buttons = browser.find_elements(By.XPATH, "//compose/ul/li/button")

    storesList = [list(stores) for stores in zip(storeText, buttons)]

    storeResults = {}

    for index, store in enumerate(storesList):
        if store[0] in STORES_TO_SCRAPE:
                
            wait_for_store_list(store[0])
            time.sleep(2)
            store[1] = browser.find_elements(By.XPATH, "//compose/ul/li/button")[index]
            print("after resetting button")
            button = store[1]
            button.location_once_scrolled_into_view
            button.click()
            WebDriverWait(browser, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//a[@click.delegate='showFulfillmentTimes($event)']")))
            
            try:
                if browser.find_element(By.XPATH, "//a[@click.delegate='showFulfillmentTimes($event)']").is_displayed():
                    browser.find_element(By.XPATH, "//a[@click.delegate='showFulfillmentTimes($event)']").click()
                    time.sleep(0.5)
            except:
                pass
            
            
            time_slots = browser.find_elements(By.XPATH, "//div[@class='radio-box-text-title']")
            time_slots = [slot.text.rstrip() for slot in time_slots]
            slots_left = browser.find_elements(By.XPATH, "//div[@class='radio-box-text-subtitle']/span")
            slots_left = [slot.text.rstrip() for slot in slots_left]

                    
            for i in range(len(slots_left)):
                if slots_left[i] == "3 Slots Left":
                    slots_left[i] = 0
                elif slots_left[i] == "2 Slots Left":
                    slots_left[i] = 1
                elif slots_left[i] == "1 Slot Left":
                    slots_left[i] = 2
                elif slots_left[i] == "Slot full":
                    slots_left[i] = 3
                else:
                    slots_left[i] = "Error"
            
            open_or_closed = ['' for slot in time_slots]
            offset = datetime.timedelta(hours=4)
            pm_offset = datetime.timedelta(hours=12)
    
            for index, slot in enumerate(time_slots):
                t = slot[0:5].rstrip()
                t = datetime.datetime.strptime(t, '%H:%M').time()
                target = datetime.datetime.combine(DAY, t)
                if "p" in slot:
                    if (NOW + offset) >= (target + pm_offset):
                        open_or_closed[index] = "CLOSED"
                    else:
                        open_or_closed[index] = "OPEN"
                else:
                    if (NOW + offset) >= target:
                        open_or_closed[index] = "CLOSED"
                    else:
                        open_or_closed[index] = "OPEN"
                            
            storeResults[store[0]] = [list(data) for data in zip(time_slots, slots_left, open_or_closed)]
            print(storeResults)


            # click "change store" button
            browser.find_element(By.XPATH, "//button[@click.trigger='controller.cancel()']").click()
            if browser.find_element(By.XPATH, "//a[@click.delegate='changeFulfillment($event)']").is_displayed():
                browser.find_element(By.XPATH, "//a[@click.delegate='changeFulfillment($event)']").click()
        else:
            pass


    # Excel functionality

    wb = Workbook()
    initial_sheet = wb.active
    initial_sheet.title = "Fareway Orders"

    nextRow = 2


    for key, value in storeResults.items():
        initial_sheet.cell(row=nextRow, column=1, value=key)
        nextRow += 1 
        for item in value:
            initial_sheet.cell(row=nextRow, column=2, value=item[0])
            initial_sheet.cell(row=nextRow, column=3, value=item[1])
            initial_sheet.cell(row=nextRow, column=4, value=item[2])
            nextRow += 1
        
        nextRow += 1
    filename = f'{TODAY}-{NOW.hour}{NOW.second}.xlsx'
    wb.save(filename)

    # Email functionality 

    sender_email = "" # REMOVED
    receiver_email = "" # FOR
    pw = "" # OBVIOUS REASONS

    message = MIMEMultipart("alternative")
    message["Subject"] = f"{TODAY} Fareway data "
    message["From"] = sender_email
    message["To"] = receiver_email

    body = f"""\
        Here's the latest scraping data as of {NOW}"""
    

    message.attach(MIMEText(body, "plain"))


    with open(filename, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    encoders.encode_base64(part)

    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    message.attach(part)
    text = message.as_string()
    

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, pw)
        server.sendmail(sender_email, receiver_email, text)
